#!/usr/bin/env python3
"""
Extract text from bank statements for the convert command.
Outputs JSON: {"text": "...", "pages": N, "error": null}
On failure: {"text": null, "pages": 0, "error": "..."}

Supported formats:
  .pdf       — pypdf (text) or pdf2image + pytesseract (OCR fallback)
  .png .jpg .jpeg .gif .bmp .tiff  — pytesseract (OCR)
  .xls .xlsx — openpyxl (cells to text table)
  .csv .txt  — read as-is
"""

import json
import sys
import os
from pathlib import Path


def extract(file_path: str) -> dict:
    path = Path(file_path)
    if not path.exists():
        return {"text": None, "pages": 0, "error": f"File not found: {file_path}"}

    ext = path.suffix.lower()

    try:
        if ext == ".pdf":
            return extract_pdf(path)
        elif ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff"):
            return extract_image(path)
        elif ext in (".xls", ".xlsx"):
            return extract_xls(path)
        elif ext in (".csv", ".txt"):
            text = path.read_text(encoding="utf-8", errors="replace")
            return {"text": text, "pages": 1, "error": None}
        else:
            # Treat unknown as text
            text = path.read_text(encoding="utf-8", errors="replace")
            return {"text": text, "pages": 1, "error": None}
    except ImportError as e:
        missing = str(e).split("'")[1] if "'" in str(e) else str(e)
        return {
            "text": None,
            "pages": 0,
            "error": f"Missing Python package: {missing}. Install with: pip install {missing}",
        }
    except Exception as e:
        return {"text": None, "pages": 0, "error": str(e)}


def extract_pdf(path: Path) -> dict:
    # Try pypdf first (text-based PDFs)
    try:
        import pypdf

        reader = pypdf.PdfReader(str(path))
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text and text.strip():
                pages.append(text.strip())
        if pages and any(len(p) > 50 for p in pages):
            return {"text": "\n\n--- Page Break ---\n\n".join(pages), "pages": len(pages), "error": None}
    except ImportError:
        pass

    # Fallback: OCR via pdf2image + pytesseract
    try:
        import pdf2image
        import pytesseract

        images = pdf2image.convert_from_path(str(path), dpi=300)
        pages = []
        for img in images:
            text = pytesseract.image_to_string(img, lang="eng")
            pages.append(text.strip())
        combined = "\n\n--- Page Break ---\n\n".join(pages)
        return {"text": combined, "pages": len(pages), "error": None}
    except ImportError as e:
        raise e


def extract_image(path: Path) -> dict:
    import pytesseract
    from PIL import Image

    img = Image.open(str(path))
    text = pytesseract.image_to_string(img, lang="eng")
    return {"text": text.strip(), "pages": 1, "error": None}


def extract_xls(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            line = ",".join("" if v is None else str(v) for v in row)
            rows.append(line)
    wb.close()
    return {"text": "\n".join(rows), "pages": len(wb.sheetnames), "error": None}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"text": None, "pages": 0, "error": "Usage: extract_text.py <file-path>"}))
        sys.exit(1)

    result = extract(sys.argv[1])
    print(json.dumps(result))
    if result["error"]:
        sys.exit(1)
